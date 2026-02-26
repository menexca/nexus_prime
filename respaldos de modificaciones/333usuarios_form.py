import sys
import os
import psycopg2
import hashlib
import re
from datetime import datetime
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, 
    QPushButton, QCheckBox, QComboBox, QMessageBox, QListWidget, 
    QFrame, QApplication, QFormLayout, QTreeWidget, QTreeWidgetItem,
    QHeaderView, QTreeWidgetItemIterator, QMenu
)
from PyQt6.QtGui import QFont, QPalette, QColor, QAction
from PyQt6.QtCore import Qt
from db_config import DB_PARAMS

# --- IMPORTACIONES PARA REPORTES ---
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class UsuariosForm(QWidget):
    def __init__(self, id_usuario_actual):
        super().__init__()
        self.id_usuario_actual = id_usuario_actual
        self.rol_usuario_sesion = "Operador"
        
        self.setWindowTitle("Gestión de Usuarios y Perfiles de Acceso")
        self.resize(1200, 750)
        
        self.id_usuario_seleccionado = None
        
        self.verificar_permisos()
        self.apply_styles()
        self.init_ui()
        self.cargar_lista_usuarios()

    def verificar_permisos(self):
        try:
            conn = psycopg2.connect(**DB_PARAMS)
            cur = conn.cursor()
            cur.execute("SELECT rol FROM seg_usuarios WHERE id_usuario = %s", (self.id_usuario_actual,))
            res = cur.fetchone()
            conn.close()
            if res: self.rol_usuario_sesion = res[0]
        except Exception as e:
            print(f"Error permisos: {e}")

    def apply_styles(self):
        palette = self.palette()
        palette.setColor(QPalette.ColorRole.Window, QColor("#F0F2F5"))
        self.setPalette(palette)
        self.setFont(QFont("Segoe UI", 10))
        
        self.setStyleSheet("""
            QPushButton {
                background-color: #007BFF; color: white; border-radius: 5px;
                padding: 8px 15px; border: none; font-weight: bold;
            }
            QPushButton:hover { background-color: #0056b3; }
            QPushButton:disabled { background-color: #cccccc; color: #666666; }
            
            QPushButton#btn_reporte {
                background-color: #17A2B8; 
                padding-right: 25px; /* Espacio para la flecha del menú */
            }
            QPushButton#btn_reporte:hover {
                background-color: #138496;
            }
            QPushButton#btn_reporte::menu-indicator {
                subcontrol-origin: padding;
                subcontrol-position: right center;
            }
            
            QLineEdit, QComboBox {
                border: 1px solid #cccccc; border-radius: 4px; padding: 5px; background-color: white;
            }
            QLineEdit:focus { border: 1px solid #007BFF; }
            QLineEdit:read-only { background-color: #E0E0E0; color: #555; }
            
            QListWidget, QTreeWidget {
                border: 1px solid #cccccc; border-radius: 4px; padding: 5px; font-size: 10pt; background-color: white;
            }
            QListWidget::item { padding: 8px; }
            QListWidget::item:selected { background-color: #E6F3FF; color: #0056b3; border-left: 4px solid #007BFF; }
            
            QTreeWidget::item { padding: 4px; }
            QHeaderView::section { background-color: #E0E6ED; padding: 6px; border: none; font-weight: bold; }
        """)

    def init_ui(self):
        main_layout = QHBoxLayout()

        # --- PANEL IZQUIERDO ---
        left_panel = QVBoxLayout()
        left_panel.addWidget(QLabel("👥 Lista de Usuarios"))
        
        self.lista_usuarios = QListWidget()
        self.lista_usuarios.itemClicked.connect(self.cargar_usuario)
        self.lista_usuarios.setMaximumHeight(200)
        left_panel.addWidget(self.lista_usuarios)
        
        self.btn_nuevo = QPushButton("+ Crear Nuevo Usuario")
        self.btn_nuevo.clicked.connect(self.limpiar_formulario)
        left_panel.addWidget(self.btn_nuevo)
        
        left_panel.addSpacing(15)
        left_panel.addWidget(QLabel("📝 Datos de Identificación"))
        
        form_layout = QFormLayout()
        
        self.txt_login = QLineEdit()
        self.txt_login.setPlaceholderText("Ej: jperez")
        form_layout.addRow("Usuario (Login) *:", self.txt_login)
        
        self.txt_nombre = QLineEdit()
        self.txt_nombre.setPlaceholderText("Nombre y Apellido")
        form_layout.addRow("Nombre Completo *:", self.txt_nombre)
        
        self.txt_email = QLineEdit()
        self.txt_email.setPlaceholderText("ejemplo@correo.com")
        form_layout.addRow("Correo Electrónico:", self.txt_email)
        
        self.cmb_rol = QComboBox()
        self.cmb_rol.addItems(["Operador", "Supervisor", "Administrador"])
        form_layout.addRow("Rol General:", self.cmb_rol)
        
        self.chk_activo = QCheckBox("Usuario Activo")
        self.chk_activo.setChecked(True)
        form_layout.addRow("Estado:", self.chk_activo)
        
        left_panel.addLayout(form_layout)
        
        # --- SECCIÓN CONTRASEÑA ---
        gb_pass = QFrame()
        gb_pass.setStyleSheet("border: 1px solid #ddd; border-radius: 4px; margin-top: 10px; padding: 5px;")
        layout_pass = QFormLayout(gb_pass)
        layout_pass.addRow(QLabel("<b>Seguridad</b>"))
        
        self.txt_pass1 = QLineEdit()
        self.txt_pass1.setEchoMode(QLineEdit.EchoMode.Password)
        self.txt_pass1.setPlaceholderText("Nueva contraseña")
        layout_pass.addRow("Contraseña:", self.txt_pass1)
        
        self.txt_pass2 = QLineEdit()
        self.txt_pass2.setEchoMode(QLineEdit.EchoMode.Password)
        self.txt_pass2.setPlaceholderText("Repetir contraseña")
        layout_pass.addRow("Repetir:", self.txt_pass2)
        
        left_panel.addWidget(gb_pass)
        
        self.lbl_audit = QLabel("Creado: - | Modif: -")
        self.lbl_audit.setStyleSheet("color: gray; font-size: 10px; margin-top: 5px;")
        left_panel.addWidget(self.lbl_audit)
        
        left_panel.addStretch()
        left_widget = QWidget()
        left_widget.setLayout(left_panel)
        left_widget.setFixedWidth(380)
        main_layout.addWidget(left_widget)

        # --- PANEL DERECHO ---
        right_panel = QVBoxLayout()
        
        header_perm = QHBoxLayout()
        header_perm.addWidget(QLabel("🛡️ Permisología del Sistema"))
        header_perm.addStretch()
        
        btn_all = QPushButton("Marcar Todo")
        btn_all.setFixedSize(90, 25)
        btn_all.setStyleSheet("font-size: 10px; background-color: #6c757d;")
        btn_all.clicked.connect(lambda: self.marcar_arbol(True))
        
        btn_none = QPushButton("Desmarcar")
        btn_none.setFixedSize(80, 25)
        btn_none.setStyleSheet("font-size: 10px; background-color: #6c757d;")
        btn_none.clicked.connect(lambda: self.marcar_arbol(False))

        header_perm.addWidget(btn_all)
        header_perm.addWidget(btn_none)
        right_panel.addLayout(header_perm)

        self.tree_permisos = QTreeWidget()
        self.tree_permisos.setHeaderLabels(["Módulo / Sección", "Ver", "Crear", "Editar", "Eliminar"])
        self.tree_permisos.setColumnWidth(0, 220)
        for i in range(1, 5): self.tree_permisos.setColumnWidth(i, 70)
            
        right_panel.addWidget(self.tree_permisos)

        btns_layout = QHBoxLayout()
        
        # --- BOTÓN DE REPORTE CON MENÚ DESPLEGABLE ---
        self.btn_reporte = QPushButton("📄 Generar Reporte ▾")
        self.btn_reporte.setObjectName("btn_reporte")
        self.btn_reporte.setFixedHeight(45)
        
        # Crear el menú para elegir formato
        menu_reportes = QMenu(self)
        menu_reportes.setStyleSheet("QMenu { background-color: white; border: 1px solid #ccc; } QMenu::item { padding: 8px 25px; } QMenu::item:selected { background-color: #007BFF; color: white; }")
        
        accion_pdf = menu_reportes.addAction("📕 Exportar a PDF")
        accion_excel = menu_reportes.addAction("📗 Exportar a Excel")
        accion_txt = menu_reportes.addAction("📓 Exportar a TXT")
        
        accion_pdf.triggered.connect(lambda: self.preparar_reporte("PDF"))
        accion_excel.triggered.connect(lambda: self.preparar_reporte("EXCEL"))
        accion_txt.triggered.connect(lambda: self.preparar_reporte("TXT"))
        
        self.btn_reporte.setMenu(menu_reportes)
        btns_layout.addWidget(self.btn_reporte)
        
        btns_layout.addStretch()
        
        self.btn_eliminar = QPushButton("Eliminar Usuario")
        self.btn_eliminar.setStyleSheet("background-color: #d9534f; color: white;")
        self.btn_eliminar.clicked.connect(self.eliminar_usuario)
        btns_layout.addWidget(self.btn_eliminar)
        
        self.btn_guardar = QPushButton("GUARDAR CAMBIOS")
        self.btn_guardar.setFixedHeight(45)
        self.btn_guardar.setStyleSheet("background-color: #28a745; color: white; font-weight: bold;")
        self.btn_guardar.clicked.connect(self.guardar_usuario)
        btns_layout.addWidget(self.btn_guardar)

        if self.rol_usuario_sesion != "Administrador":
            self.btn_guardar.setEnabled(False)
            self.btn_eliminar.setEnabled(False)
            self.btn_nuevo.setEnabled(False)

        right_panel.addLayout(btns_layout)
        main_layout.addLayout(right_panel)
        self.setLayout(main_layout)

    # --- LÓGICA DE NEGOCIO ---

    def hash_password(self, password):
        return hashlib.sha256(password.encode()).hexdigest()
    
    def validar_email(self, email):
        pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
        return re.match(pattern, email) is not None

    def cargar_lista_usuarios(self):
        self.lista_usuarios.clear()
        try:
            conn = psycopg2.connect(**DB_PARAMS)
            cur = conn.cursor()
            cur.execute("SELECT id_usuario, usuario_login, nombre_completo FROM seg_usuarios ORDER BY id_usuario")
            for u in cur.fetchall():
                self.lista_usuarios.addItem(f"{u[1]} - {u[2]}")
                self.lista_usuarios.item(self.lista_usuarios.count()-1).setData(Qt.ItemDataRole.UserRole, u[0])
            conn.close()
            self.cargar_estructura_arbol()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error cargando lista: {str(e)}")

    def create_checkbox_widget(self):
        widget = QWidget()
        chk = QCheckBox()
        layout = QHBoxLayout(widget)
        layout.addWidget(chk)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.setContentsMargins(0, 0, 0, 0)
        return widget, chk

    def cargar_estructura_arbol(self):
        self.tree_permisos.clear()
        try:
            conn = psycopg2.connect(**DB_PARAMS)
            cur = conn.cursor()
            cur.execute("SELECT id_modulo, nombre_modulo, categoria FROM sys_moduser ORDER BY categoria, nombre_modulo")
            modulos = cur.fetchall()
            conn.close()

            categorias = {}
            for m in modulos:
                cat = m[2]
                if cat not in categorias:
                    parent = QTreeWidgetItem(self.tree_permisos)
                    parent.setText(0, cat.upper())
                    parent.setExpanded(True)
                    for c in range(5): parent.setBackground(c, QColor("#F2F2F2"))
                    categorias[cat] = parent
                
                child = QTreeWidgetItem(categorias[cat])
                child.setText(0, m[1])
                child.setData(0, Qt.ItemDataRole.UserRole, m[0]) 
                
                for col in range(1, 5):
                    container, chk = self.create_checkbox_widget()
                    self.tree_permisos.setItemWidget(child, col, container)
                    
        except Exception as e:
            print(f"Error cargando árbol: {e}")

    def cargar_usuario(self, item):
        uid = item.data(Qt.ItemDataRole.UserRole)
        self.id_usuario_seleccionado = uid
        self.cargar_estructura_arbol() 

        try:
            conn = psycopg2.connect(**DB_PARAMS)
            cur = conn.cursor()
            
            query = """
                SELECT u.usuario_login, u.nombre_completo, u.email, u.rol, u.estatus, 
                       u1.usuario_login, u.fecha_creacion, 
                       u2.usuario_login, u.fecha_modifica
                FROM seg_usuarios u
                LEFT JOIN seg_usuarios u1 ON u.id_user_crea = u1.id_usuario
                LEFT JOIN seg_usuarios u2 ON u.id_user_mod = u2.id_usuario
                WHERE u.id_usuario = %s
            """
            cur.execute(query, (uid,))
            data = cur.fetchone()
            
            if data:
                self.txt_login.setText(data[0]); self.txt_login.setReadOnly(True)
                self.txt_nombre.setText(data[1])
                self.txt_email.setText(data[2] if data[2] else "")
                self.cmb_rol.setCurrentText(data[3] or "Operador")
                self.chk_activo.setChecked(data[4])
                
                f_crea = str(data[6])[:16] if data[6] else "-"
                f_mod = str(data[8])[:16] if data[8] else "-"
                self.lbl_audit.setText(f"Crea: {data[5] or '-'} ({f_crea}) | Mod: {data[7] or '-'} ({f_mod})")
                self.txt_pass1.clear(); self.txt_pass2.clear()

            cur.execute("SELECT id_modulo, p_ver, p_crear, p_editar, p_eliminar FROM sys_permisouser WHERE id_usuario = %s", (uid,))
            permisos_db = cur.fetchall()
            
            mapa_permisos = { row[0]: row[1:] for row in permisos_db }

            iterator = QTreeWidgetItemIterator(self.tree_permisos)
            while iterator.value():
                item_tree = iterator.value()
                id_mod = item_tree.data(0, Qt.ItemDataRole.UserRole)
                
                if id_mod and id_mod in mapa_permisos:
                    vals = mapa_permisos[id_mod]
                    for i in range(4):
                        col_idx = i + 1
                        container = self.tree_permisos.itemWidget(item_tree, col_idx)
                        if container:
                            chk = container.findChild(QCheckBox)
                            if chk: chk.setChecked(vals[i])
                iterator += 1
            conn.close()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error cargando usuario: {str(e)}")

    def guardar_usuario(self):
        if self.rol_usuario_sesion != "Administrador": return
        
        login = self.txt_login.text().strip()
        nombre = self.txt_nombre.text().strip()
        email = self.txt_email.text().strip()
        p1 = self.txt_pass1.text()
        p2 = self.txt_pass2.text()
        
        if not login or not nombre: 
            QMessageBox.warning(self, "Datos incompletos", "El Usuario y Nombre Completo son obligatorios.")
            return

        if email and not self.validar_email(email):
            QMessageBox.warning(self, "Email inválido", "El formato del correo no es correcto.")
            return

        pass_hash = None
        
        if self.id_usuario_seleccionado is None:
            if not p1: 
                QMessageBox.warning(self, "Seguridad", "Debe asignar una contraseña al nuevo usuario.")
                return
            if p1 != p2:
                QMessageBox.warning(self, "Error", "Las contraseñas no coinciden.")
                return
            pass_hash = self.hash_password(p1)
        else:
            if p1: 
                if p1 != p2:
                    QMessageBox.warning(self, "Error", "Las contraseñas no coinciden.")
                    return
                pass_hash = self.hash_password(p1)

        try:
            conn = psycopg2.connect(**DB_PARAMS)
            cur = conn.cursor()
            new_id = self.id_usuario_seleccionado
            
            if new_id is None:
                cur.execute("SELECT 1 FROM seg_usuarios WHERE usuario_login=%s", (login,))
                if cur.fetchone(): 
                    QMessageBox.warning(self, "Duplicado", f"El usuario '{login}' ya existe.")
                    conn.close()
                    return
                
                cur.execute("""
                    INSERT INTO seg_usuarios (usuario_login, nombre_completo, email, password_hash, rol, estatus, id_user_crea)
                    VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id_usuario
                """, (login, nombre, email, pass_hash, self.cmb_rol.currentText(), self.chk_activo.isChecked(), self.id_usuario_actual))
                new_id = cur.fetchone()[0]
            else:
                if pass_hash:
                    cur.execute("""
                        UPDATE seg_usuarios SET nombre_completo=%s, email=%s, password_hash=%s, rol=%s, estatus=%s, id_user_mod=%s
                        WHERE id_usuario=%s
                    """, (nombre, email, pass_hash, self.cmb_rol.currentText(), self.chk_activo.isChecked(), self.id_usuario_actual, new_id))
                else:
                    cur.execute("""
                        UPDATE seg_usuarios SET nombre_completo=%s, email=%s, rol=%s, estatus=%s, id_user_mod=%s
                        WHERE id_usuario=%s
                    """, (nombre, email, self.cmb_rol.currentText(), self.chk_activo.isChecked(), self.id_usuario_actual, new_id))
            
            cur.execute("DELETE FROM sys_permisouser WHERE id_usuario = %s", (new_id,))
            
            iterator = QTreeWidgetItemIterator(self.tree_permisos)
            while iterator.value():
                item = iterator.value()
                id_mod = item.data(0, Qt.ItemDataRole.UserRole)
                
                if id_mod: 
                    vals = []
                    for col in range(1, 5):
                        container = self.tree_permisos.itemWidget(item, col)
                        checked = False
                        if container:
                            chk = container.findChild(QCheckBox)
                            if chk: checked = chk.isChecked()
                        vals.append(checked)
                    
                    if any(vals):
                        cur.execute("""
                            INSERT INTO sys_permisouser (id_usuario, id_modulo, p_ver, p_crear, p_editar, p_eliminar)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (new_id, id_mod, vals[0], vals[1], vals[2], vals[3]))
                
                iterator += 1

            conn.commit()
            conn.close()
            QMessageBox.information(self, "Éxito", "Usuario guardado correctamente.")
            self.limpiar_formulario()
            self.cargar_lista_usuarios()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Fallo al guardar: {e}")

    def eliminar_usuario(self):
        if self.rol_usuario_sesion != "Administrador": return
        if not self.id_usuario_seleccionado: return
        if self.id_usuario_seleccionado == self.id_usuario_actual:
            QMessageBox.warning(self, "Error", "No puedes auto-eliminarte.")
            return

        if QMessageBox.question(self, "Confirmar", "¿Eliminar usuario definitivamente?") == QMessageBox.StandardButton.Yes:
            try:
                conn = psycopg2.connect(**DB_PARAMS)
                cur = conn.cursor()
                cur.execute("DELETE FROM seg_usuarios WHERE id_usuario=%s", (self.id_usuario_seleccionado,))
                conn.commit()
                conn.close()
                self.limpiar_formulario()
                self.cargar_lista_usuarios()
            except Exception as e:
                QMessageBox.critical(self, "Error", "No se puede eliminar (posiblemente tenga registros asociados).")

    def limpiar_formulario(self):
        self.id_usuario_seleccionado = None
        self.txt_login.clear(); self.txt_login.setReadOnly(False)
        self.txt_nombre.clear()
        self.txt_email.clear()
        self.txt_pass1.clear(); self.txt_pass2.clear()
        self.lbl_audit.setText("Nuevo Usuario")
        self.cargar_estructura_arbol()

    # ==========================================
    # LÓGICA DE EXPORTACIÓN Y REPORTES
    # ==========================================
    def preparar_reporte(self, formato):
        if not self.id_usuario_seleccionado:
            QMessageBox.warning(self, "Advertencia", "Debe seleccionar un usuario de la lista primero.")
            return

        uid = self.id_usuario_seleccionado
        try:
            conn = psycopg2.connect(**DB_PARAMS)
            cur = conn.cursor()
            
            # 1. Datos del Usuario
            cur.execute("SELECT usuario_login, nombre_completo, email, rol, estatus FROM seg_usuarios WHERE id_usuario = %s", (uid,))
            usuario_data = cur.fetchone()
            
            # 2. Empresas asociadas
            cur.execute("""
                SELECT c.razon_social, c.rif 
                FROM sys_acceso_empresas a
                JOIN cfg_empresas c ON a.cod_compania = c.cod_compania
                WHERE a.id_usuario = %s
                ORDER BY c.razon_social
            """, (uid,))
            empresas_data = cur.fetchall()

            # 3. Permisos
            cur.execute("""
                SELECT m.categoria, m.nombre_modulo, p.p_ver, p.p_crear, p.p_editar, p.p_eliminar
                FROM sys_permisouser p
                JOIN sys_moduser m ON p.id_modulo = m.id_modulo
                WHERE p.id_usuario = %s
                ORDER BY m.categoria, m.nombre_modulo
            """, (uid,))
            permisos_data = cur.fetchall()
            conn.close()

            # Enrutar según formato elegido
            if formato == "PDF":
                self.generar_pdf(usuario_data, empresas_data, permisos_data)
            elif formato == "EXCEL":
                self.generar_excel(usuario_data, empresas_data, permisos_data)
            elif formato == "TXT":
                self.generar_txt(usuario_data, empresas_data, permisos_data)
                
        except Exception as e:
            QMessageBox.critical(self, "Error BD", f"Error al recopilar datos para el reporte: {str(e)}")

    def abrir_archivo(self, filepath):
        """Abre el archivo generado en el programa predeterminado de Windows/Mac/Linux"""
        if sys.platform == "win32":
            os.startfile(filepath)
        elif sys.platform == "darwin":
            import subprocess
            subprocess.call(["open", filepath])
        else:
            import subprocess
            subprocess.call(["xdg-open", filepath])

    def generar_pdf(self, u_data, e_data, p_data):
        pdf_filename = f"Reporte_Usuario_{u_data[0]}.pdf"
        doc = SimpleDocTemplate(pdf_filename, pagesize=letter)
        elementos = []
        estilos = getSampleStyleSheet()
        
        estilo_titulo = estilos['Heading1']
        estilo_titulo.alignment = 1 
        estilo_subtitulo = estilos['Heading2']
        estilo_texto = estilos['Normal']

        elementos.append(Paragraph("NEXUS ERP - Reporte de Seguridad", estilo_titulo))
        elementos.append(Paragraph(f"Fecha de Emisión: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", estilo_texto))
        elementos.append(Spacer(1, 20))

        elementos.append(Paragraph("Datos de Identificación del Usuario", estilo_subtitulo))
        info_usuario = [
            ["Login:", u_data[0]], ["Nombre Completo:", u_data[1]],
            ["Correo Electrónico:", u_data[2] if u_data[2] else "N/A"],
            ["Rol en Sistema:", u_data[3]], ["Estatus Actual:", "Activo" if u_data[4] else "Inactivo"]
        ]
        t_usr = Table(info_usuario, colWidths=[150, 300])
        t_usr.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
            ('BOX', (0,0), (-1,-1), 0.25, colors.black),
            ('PADDING', (0,0), (-1,-1), 5)
        ]))
        elementos.append(t_usr)
        elementos.append(Spacer(1, 20))

        elementos.append(Paragraph("Empresas Autorizadas", estilo_subtitulo))
        if e_data:
            d_emp = [["Razón Social de la Empresa", "RIF"]]
            for emp in e_data: d_emp.append([emp[0], emp[1]])
            t_emp = Table(d_emp, colWidths=[300, 150])
            t_emp.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#007BFF')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                ('BOX', (0,0), (-1,-1), 0.25, colors.black),
            ]))
            elementos.append(t_emp)
        else:
            elementos.append(Paragraph("No tiene acceso a ninguna empresa.", estilo_texto))
        
        elementos.append(Spacer(1, 20))

        elementos.append(Paragraph("Módulos y Permisos Asignados", estilo_subtitulo))
        if p_data:
            d_perm = [["Categoría", "Módulo", "Ver", "Crear", "Editar", "Eliminar"]]
            for p in p_data:
                d_perm.append([p[0], p[1], "Sí" if p[2] else "No", "Sí" if p[3] else "No", "Sí" if p[4] else "No", "Sí" if p[5] else "No"])
            t_perm = Table(d_perm, colWidths=[100, 150, 50, 50, 50, 50])
            t_perm.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#28a745')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (2,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                ('BOX', (0,0), (-1,-1), 0.25, colors.black),
            ]))
            elementos.append(t_perm)
        else:
            elementos.append(Paragraph("No cuenta con privilegios en módulos.", estilo_texto))

        doc.build(elementos)
        self.abrir_archivo(pdf_filename)

    def generar_excel(self, u_data, e_data, p_data):
        excel_filename = f"Reporte_Usuario_{u_data[0]}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Seguridad"

        # Estilos Excel
        ft_titulo = Font(size=14, bold=True, color="FFFFFF")
        fill_titulo = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        ft_negrita = Font(bold=True)
        fill_cabecera = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Escribir Datos
        ws.append(["NEXUS ERP - Reporte de Seguridad"])
        ws['A1'].font = ft_titulo
        ws['A1'].fill = fill_titulo
        ws.merge_cells('A1:F1')
        
        ws.append([f"Fecha de Emisión: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        ws.append([])
        
        # Bloque Usuario
        ws.append(["DATOS DEL USUARIO"])
        ws['A4'].font = ft_negrita
        ws.append(["Login:", u_data[0]])
        ws.append(["Nombre Completo:", u_data[1]])
        ws.append(["Correo Electrónico:", u_data[2] if u_data[2] else "N/A"])
        ws.append(["Rol:", u_data[3]])
        ws.append(["Estatus:", "Activo" if u_data[4] else "Inactivo"])
        for row in range(5, 10): ws[f'A{row}'].font = ft_negrita
        ws.append([])

        # Bloque Empresas
        ws.append(["EMPRESAS AUTORIZADAS"])
        r_idx = ws.max_row
        ws[f'A{r_idx}'].font = ft_negrita
        ws.append(["Razón Social", "RIF"])
        for col in ['A', 'B']: 
            ws[f'{col}{ws.max_row}'].font = ft_negrita
            ws[f'{col}{ws.max_row}'].fill = fill_cabecera
            
        if e_data:
            for emp in e_data: ws.append([emp[0], emp[1]])
        else:
            ws.append(["Sin acceso a empresas"])
        ws.append([])

        # Bloque Permisos
        ws.append(["MÓDULOS Y PERMISOS"])
        r_idx = ws.max_row
        ws[f'A{r_idx}'].font = ft_negrita
        ws.append(["Categoría", "Módulo", "Ver", "Crear", "Editar", "Eliminar"])
        for col in ['A','B','C','D','E','F']: 
            ws[f'{col}{ws.max_row}'].font = ft_negrita
            ws[f'{col}{ws.max_row}'].fill = fill_cabecera
            
        if p_data:
            for p in p_data:
                ws.append([p[0], p[1], "Sí" if p[2] else "No", "Sí" if p[3] else "No", "Sí" if p[4] else "No", "Sí" if p[5] else "No"])
        else:
            ws.append(["Sin permisos asignados"])

        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35

        wb.save(excel_filename)
        self.abrir_archivo(excel_filename)

    def generar_txt(self, u_data, e_data, p_data):
        txt_filename = f"Reporte_Usuario_{u_data[0]}.txt"
        
        with open(txt_filename, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write(" NEXUS ERP - REPORTE DE SEGURIDAD\n")
            f.write("=" * 60 + "\n")
            f.write(f"Fecha de Emisión: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n\n")
            
            f.write("--- DATOS DEL USUARIO ---\n")
            f.write(f"Login:            {u_data[0]}\n")
            f.write(f"Nombre Completo:  {u_data[1]}\n")
            f.write(f"Email:            {u_data[2] if u_data[2] else 'N/A'}\n")
            f.write(f"Rol:              {u_data[3]}\n")
            f.write(f"Estatus:          {'Activo' if u_data[4] else 'Inactivo'}\n\n")
            
            f.write("--- EMPRESAS AUTORIZADAS ---\n")
            if e_data:
                for emp in e_data:
                    f.write(f"- {emp[0]} (RIF: {emp[1]})\n")
            else:
                f.write("No tiene acceso a ninguna empresa.\n")
            f.write("\n")
            
            f.write("--- MÓDULOS Y PERMISOS ---\n")
            if p_data:
                f.write(f"{'CATEGORÍA':<15} | {'MÓDULO':<20} | VER | CREAR | EDITAR | ELIMINAR\n")
                f.write("-" * 80 + "\n")
                for p in p_data:
                    v = "Sí" if p[2] else "No"
                    c = "Sí" if p[3] else "No"
                    e = "Sí" if p[4] else "No"
                    el = "Sí" if p[5] else "No"
                    f.write(f"{p[0]:<15} | {p[1]:<20} | {v:<3} | {c:<5} | {e:<6} | {el:<8}\n")
            else:
                f.write("No cuenta con privilegios en módulos.\n")

        self.abrir_archivo(txt_filename)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = UsuariosForm(1)
    window.showMaximized() #show()
    sys.exit(app.exec())